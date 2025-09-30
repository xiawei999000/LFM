# application of LFM for MVI and PathologicalGrade prediction
from __future__ import print_function, division
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
import numpy as np
from torchvision import transforms as T
import time
import math
import os
import torch.utils.data as data
import pickle
from sklearn.metrics import roc_auc_score
from _collections import OrderedDict
import openpyxl
from msn_custom_ViT_model import MSN, vit
from monai.utils import set_determinism
import torch.nn.functional as F

def log_create():
    # INFO
    # data sheet
    info = openpyxl.Workbook()

    all_sheet = info.create_sheet('1', 0)
    all_sheet.cell(1, 1, 'learning_rate')
    all_sheet.cell(1, 2, 'batch_size')
    all_sheet.cell(1, 3, 'epoch_id')

    all_sheet.cell(1, 4, 'AUC_train_mean')
    all_sheet.cell(1, 5, 'AUC_val_mean')

    all_sheet.cell(1, 6, 'MVI_AUC_train')
    all_sheet.cell(1, 7, 'MVI_AUC_val')


    all_sheet.cell(1, 8, 'PathologicalGrade_AUC_train')
    all_sheet.cell(1, 9, 'PathologicalGrade_AUC_val')

    return info

class LN_Dataset(data.Dataset):

    def __init__(self, p_patch_nc_list, p_patch_ap_list, p_patch_pv_list, p_label_list):
        self.list_p_patch_nc = p_patch_nc_list
        self.list_p_patch_ap = p_patch_ap_list
        self.list_p_patch_pv = p_patch_pv_list
        self.list_p_label = p_label_list

    def __getitem__(self, idx):
        p_patch_nc = self.list_p_patch_nc[idx]
        p_patch_ap = self.list_p_patch_ap[idx]
        p_patch_pv = self.list_p_patch_pv[idx]
        p_label = self.list_p_label[idx]
        return p_patch_nc, p_patch_ap, p_patch_pv, p_label

    def __len__(self):
        return len(self.list_p_label)

def my_collate(batch):
    p_patches_nc = []
    p_patches_ap = []
    p_patches_pv = []
    p_labels = []

    for p_patch_nc, p_patch_ap, p_patch_pv, p_label in batch:
        p_patches_nc.append(p_patch_nc)
        p_patches_ap.append(p_patch_ap)
        p_patches_pv.append(p_patch_pv)
        p_labels.append(p_label)
    return p_patches_nc, p_patches_ap, p_patches_pv, p_labels


def evaluate(y_true, y_pred, digits=3):    
    AUC = round(roc_auc_score(y_true, y_pred), digits)
    return AUC


def split_data(images, labels, set_split):

    X_train = [img for img, set_name in zip(images, set_split) if set_name == 'train']
    y_train = [label for label, set_name in zip(labels, set_split) if set_name == 'train']

    X_val = [img for img, set_name in zip(images, set_split) if set_name == 'val']
    y_val = [label for label, set_name in zip(labels, set_split) if set_name == 'val']

    return X_train, X_val, y_train, y_val


class Patches_ABMIL(nn.Module):
    def __init__(self, vit_model, feature_num):
        super(Patches_ABMIL, self).__init__()
        self.L = 128

        self.transformer = vit_model.backbone

        self.attention_V = nn.Sequential(
            nn.Linear(feature_num, self.L),  # matrix V
            nn.Tanh()
        )

        self.attention_U = nn.Sequential(
            nn.Linear(feature_num, self.L),  # matrix U
            nn.Sigmoid()
        )

        self.attention_w = nn.Linear(self.L, 1)

    def forward(self, p_patches):
        patch_num = len(p_patches)

        p_embeddings = []
        p_embeddings = torch.tensor(p_embeddings)
        p_embeddings = p_embeddings.to(device)

        for patch_id in range(0, patch_num):
            p_patch = p_patches[patch_id]

            p_patch = T.ToTensor()(p_patch)
            p_patch = p_patch.unsqueeze(0)
            p_patch = torch.as_tensor(p_patch, dtype=torch.float)
            p_patch = p_patch.to(device)

            p_patch_embedding = self.transformer(images=p_patch)
            p_embeddings = torch.cat((p_embeddings, p_patch_embedding))

        A_V = self.attention_V(p_embeddings)  # KxL
        A_U = self.attention_U(p_embeddings)  # KxL
        A = self.attention_w(A_V * A_U)  # element wise multiplication # KxATTENTION_BRANCHES
        A = torch.transpose(A, 1, 0)  # ATTENTION_BRANCHESxK
        A = F.softmax(A, dim=1)  # softmax over K

        Z = torch.mm(A, p_embeddings)

        Z = Z.squeeze()

        return Z

class FusionClassifier(nn.Module):
    def __init__(self, model_nc, model_ap, model_pv, feature_num, num_classes):
        super(FusionClassifier, self).__init__()
     
        self.branch_nc = model_nc
        self.branch_ap = model_ap
        self.branch_pv = model_pv

        self.feas_mlp_head_nc = nn.Sequential(
            nn.Linear(feature_num, 128),
            nn.ReLU(),
        )

        self.feas_mlp_head_ap = nn.Sequential(
            nn.Linear(feature_num, 128),
            nn.ReLU(),
        )

        self.feas_mlp_head_pv = nn.Sequential(
            nn.Linear(feature_num, 128),
            nn.ReLU(),
        )

        self.combine_mlp_head = nn.Sequential(
            nn.Linear(384, 128),
            nn.ReLU(),
            nn.Linear(128, num_classes),
            nn.Sigmoid(),
        )

    def forward(self, p_patches_nc, p_patches_ap, p_patches_pv):

        features_nc = self.branch_nc(p_patches_nc)
        features_nc =self.feas_mlp_head_nc(features_nc)

        features_ap = self.branch_ap(p_patches_ap)
        features_ap = self.feas_mlp_head_ap(features_ap)

        features_pv = self.branch_pv(p_patches_pv)
        features_pv = self.feas_mlp_head_pv(features_pv)

        # combine the features of three phases
        p_prediction_combine = self.combine_mlp_head(
            torch.cat((features_nc, features_ap, features_pv)))

        return p_prediction_combine


def early_stopping(val_loss, patience=10):
    if len(val_loss) > 1 and val_loss[-1] <= np.array(val_loss[:-1]).min():
        return False

    if len(val_loss) > patience and val_loss[-1] > val_loss[-patience - 1]:
        print(f'Early stopping at epoch {epoch}')
        return True
    return False

def train_model(device, dataloaders, model, criterions, optimizer, scheduler,
                batch_size, learning_rate, num_epochs, work_dir, log_path, log_modelling, index_record, label_obj_list):

    n_class = len(label_obj_list)

    since = time.time()
    # online data augmentation in training
    random_transform = T.Compose([
        # 1 Horizontal Flip
        T.RandomHorizontalFlip(),
        # 2 Vertical Flip
        T.RandomVerticalFlip(),
        # 3 RandomRotation
        T.RandomRotation(10),
    ])

    # folder to store models, logs
    # folder named by hyper-paramters
    results_save_folder = work_dir + 'lr_{}_bs_{}/'.format(learning_rate, batch_size)
    if not os.path.exists(results_save_folder):
        os.makedirs(results_save_folder)

    # the sheet to store infomation
    shenames = log_modelling.get_sheet_names()
    all_sheet = log_modelling[shenames[0]]
    #
    AUC_train = 0.0
    AUC_val = 0.0

    best_val_auc = 0.0

    val_loss = []
    patience = 5
    outer_break = False


    for epoch in range(num_epochs):
        epoch_id = epoch + 1
        print('Epoch {}/{}'.format(epoch_id, num_epochs))
        print('-' * 10)
        print('running learning rate: ', optimizer.param_groups[0]['lr'])

        epoch_start_time = time.time()

        # write log
        all_sheet.cell(index_record, 1, learning_rate) 
        all_sheet.cell(index_record, 2, batch_size)
        all_sheet.cell(index_record, 3, epoch_id)

        for phase in ['train', 'val']:  #
            # phase_flag = 0
            if phase == 'train':
                model.train()  # Set model to training mode
                phase_flag = 0
            elif phase == 'val':
                model.eval()   # Set model to evaluate mode
                phase_flag = 1

            torch.set_grad_enabled(phase == 'train')

            size = len(dataloaders[phase_flag].dataset)
            num_mini_batches = math.ceil(size / batch_size)
            print('num_mini_batches = ', num_mini_batches)

            running_loss = 0.0

            epoch_outputs_p = []
            epoch_labels_p = []

            # Iterate over data.
            for batch, (p_patches_nc_batch, p_patches_ap_batch, p_patches_pv_batch, p_label_batch) \
                    in enumerate(dataloaders[phase_flag]):

                # labels
                p_label_batch = np.array(p_label_batch)
                label_p_batch = torch.tensor(p_label_batch, dtype=torch.float)
                label_p_batch = label_p_batch.to(device)

                # store model outputs for a batch of patients
                outputs_p_batch = []

                # load per patient's patches
                for p_id in range(0, len(p_patches_nc_batch)):
                    p_patches_nc = p_patches_nc_batch[p_id]
                    p_patches_ap = p_patches_ap_batch[p_id]
                    p_patches_pv = p_patches_pv_batch[p_id]

                    # forward
                    patch_num = len(p_patches_nc)
                    for patch_id in range(0, patch_num):
                        # extract patches and input into model
                        p_patch_nc = p_patches_nc[patch_id]
                        p_patch_nc = T.ToTensor()(p_patch_nc)
                        p_patch_nc = p_patch_nc.permute(1, 2, 0)
                        p_patches_nc[patch_id] = p_patch_nc

                        p_patch_ap = p_patches_ap[patch_id]
                        p_patch_ap = T.ToTensor()(p_patch_ap)
                        p_patch_ap = p_patch_ap.permute(1, 2, 0)
                        p_patches_ap[patch_id] = p_patch_ap

                        p_patch_pv = p_patches_pv[patch_id]
                        p_patch_pv = T.ToTensor()(p_patch_pv)
                        p_patch_pv = p_patch_pv.permute(1, 2, 0)
                        p_patches_pv[patch_id] = p_patch_pv

                    p_pred = model(p_patches_nc, p_patches_ap, p_patches_pv)

                    outputs_p_batch.append(p_pred)

                outputs_p_batch = torch.stack(outputs_p_batch)
                loss = criterions['loss_p_combine'](outputs_p_batch, label_p_batch)

                if phase == 'train':
                    optimizer.zero_grad()
                    loss.backward()
                    optimizer.step()
                # print internal information
                loss_item, current = loss.item(), batch * batch_size
                print(f"batch {batch+1} - loss: {loss_item:>4f}  [{current:>5d}/{size:>5d}]")

                running_loss += loss.item()  # total loss for one epoch
                
                # prepare for auc by storing each batch
                epoch_outputs_p.extend(outputs_p_batch.tolist())
                epoch_labels_p.extend(p_label_batch)

            # statistics of one epoch
            epoch_loss = running_loss / num_mini_batches

            epoch_labels_p = np.array(epoch_labels_p)

            epoch_outputs_p = np.array(epoch_outputs_p)

            auc_scores = []
            for i in range(n_class):
                label_i = epoch_labels_p[:, i]
                label_pred_i = epoch_outputs_p[:, i]
                auc = roc_auc_score(label_i, label_pred_i)
                auc_scores.append(auc)
            mean_auc = np.mean(auc_scores)

            print(
                '{} Loss: {:.4f} AUC_mean:{:.3f} MVI: {:.3f} PathologicalGrade: {:.3f} '.format(
                phase, epoch_loss, mean_auc,
                    auc_scores[0], auc_scores[1]))

            # record the train val and test performances
            if phase == 'train':
                all_sheet.cell(index_record, 4, mean_auc)
                all_sheet.cell(index_record, 6, auc_scores[0])
                all_sheet.cell(index_record, 8, auc_scores[1])
                AUC_train = mean_auc

            if phase == 'val':
                all_sheet.cell(index_record, 5, mean_auc)
                all_sheet.cell(index_record, 7, auc_scores[0])
                all_sheet.cell(index_record, 9, auc_scores[1])
                AUC_val = mean_auc

                # early stopping
                val_loss.append(epoch_loss)
                scheduler.step()

                early_stop = early_stopping(val_loss, patience=patience)
                if early_stop:
                    outer_break = True
                    break

        if outer_break:
            break

        # save the best models
        if AUC_val > best_val_auc:
            best_val_auc = AUC_val
            torch.save(model.state_dict(),
                       results_save_folder +
                       'model_mean_auc_{:.3f}_{:.3f}_epoch_{}.pkl'.format(AUC_train, AUC_val, epoch + 1))
            print("saved new best model in val")

        # write log for the next epoch
        index_record = index_record + 1

        # computation time of each epoch
        epoch_end_time = time.time()
        print('epoch computation time: ', str(epoch_end_time-epoch_start_time))
    # save the training log
    log_modelling.save(filename=log_path)
    # computation time for entire training
    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('best_val_auc: {:.4f}'.format(best_val_auc))
    print()
    print()

    return index_record


if __name__ == '__main__':
    # ensure the reproducbility
    date = 241104
    set_determinism()

    # data folder
    data_folder_path = '/data/HCC-bin/'

    result_folder_path = data_folder_path + 'results_FM/'

    if not os.path.exists(result_folder_path):
        os.makedirs(result_folder_path)

    ################# model training ########################
    # initialize the GPU
    device = torch.device("cuda:0")
    # set hyper-parameters for model training
    batch_size = 16
    learning_rate = 2e-3
    epoch = 50

    # phases : non-contrast arterial-phase portal-venous
    phase_list = ['arterial-phase', 'portal-venous', 'non-contrast']

    label_obj_list = ['MVI', 'PathologicalGrade']
    # 读取每个标签数据
    labels_dict = {}
    for label_name in label_obj_list:
        label_path = os.path.join(data_folder_path, 'HCC_' + label_name + '.bin')

        labels_dict[label_name] = pickle.load(open(label_path, "rb"))

    labels_array = np.column_stack([labels_dict[label] for label in labels_dict])

    ######################## load training and validation data  ########################
    # load data
    p_patches_nc = pickle.load(
        open(data_folder_path + '/HCC_patches_non-contrast.bin', "rb"))
    p_patches_ap = pickle.load(
        open(data_folder_path + '/HCC_patches_arterial-phase.bin', "rb"))
    p_patches_pv = pickle.load(
        open(data_folder_path + '/HCC_patches_portal-venous.bin', "rb"))

    set_split = pickle.load(open(data_folder_path + '/set_split.bin', "rb"))

    p_patches_nc_train, p_patches_nc_val, p_label_train, p_label_val = split_data(
        p_patches_nc, labels_array, set_split)
    p_patches_ap_train, p_patches_ap_val, _, _, _ = split_data(
        p_patches_ap, labels_array, set_split)
    p_patches_pv_train, p_patches_pv_val, _, _, _ = split_data(
        p_patches_pv, labels_array, set_split)

    print("Training set size:", len(p_label_train))
    print("Validation set size:", len(p_label_val))

    model_name = 'FusionPhase_FM_ABMIL'

    # path to store results
    work_dir = result_folder_path + model_name + '/'
    if not os.path.exists(work_dir):
        os.makedirs(work_dir)
    log_path = work_dir + '/log.xlsx'
    log_modelling = log_create()
    log_modelling.save(filename=log_path)
    index_record = 2


    print(f"batch_size: {batch_size}; learning_rate: {learning_rate}")
    print(
        "Number of samples in train, validation are %d, %d."
        % (len(p_label_train), len(p_label_val)))

    # load pytorch dataset
    train_dataset = LN_Dataset(p_patch_nc_list=p_patches_nc_train,
                               p_patch_ap_list=p_patches_ap_train,
                               p_patch_pv_list=p_patches_pv_train,
                               p_label_list=p_label_train
                               )
    train_dataloader = torch.utils.data.DataLoader(train_dataset, batch_size=batch_size, shuffle=True,
                                                   num_workers=16, pin_memory=True, collate_fn=my_collate)

    val_dataset = LN_Dataset(p_patch_nc_list=p_patches_nc_val,
                             p_patch_ap_list=p_patches_ap_val,
                             p_patch_pv_list=p_patches_pv_val,
                             p_label_list=p_label_val
                             )
    val_dataloader = torch.utils.data.DataLoader(val_dataset, batch_size=batch_size,
                                                 num_workers=16, pin_memory=True, collate_fn=my_collate)

    dataloaders = [train_dataloader, val_dataloader]

    model = MSN(vit)

    # load pretrained model
    
    with open('./ckps/LFM_pretraining/LFM.pth', 'rb') as f:
        model_dict = torch.load(f)
    model.load_state_dict(model_dict)

    num_classes_downstream = len(label_obj_list)
    model_nc = Patches_ABMIL(model, 512)
    model_ap = Patches_ABMIL(model, 512)
    model_pv = Patches_ABMIL(model, 512)
    model_ft = FusionClassifier(model_nc, model_ap, model_pv, 512, num_classes_downstream)

    model_ft.to(device)

    # Observe that all parameters are being optimized
    optimizer_ft = optim.AdamW(model_ft.parameters(), lr=learning_rate)

    # lr_scheduler
    CosineAnnealingLR_scheduler = lr_scheduler.CosineAnnealingLR(optimizer_ft, T_max=(epoch * len(
        p_label_train)) // batch_size / 10, eta_min=1e-8, verbose=True)

    # define loss function
    loss_fuc = 'BCELoss'
    criterions = OrderedDict()
    criterions['loss_p_combine'] = nn.BCELoss()

    torch.backends.cudnn.benchmark = True
    ind_rec = train_model(device, dataloaders, model_ft, criterions,
                optimizer_ft, CosineAnnealingLR_scheduler, batch_size,
                learning_rate, epoch, work_dir, log_path, log_modelling, index_record, label_obj_list)
    index_record = ind_rec
